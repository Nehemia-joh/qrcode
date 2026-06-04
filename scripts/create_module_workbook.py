#!/usr/bin/env python3
"""Create blank Kitchen / Farm / Facilities workbook templates for import."""

from __future__ import annotations

import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    raise SystemExit(1)

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "templates"

KPI_ROWS = [
    ("Key Performance Indicator", "Target", "Actual"),
    ("1. Sample KPI one", 90, 88),
    ("2. Sample KPI two", 100, 95),
    ("3. Sample KPI three", 0.05, 0.04),
]


def build_workbook(title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = title
    ws["C3"] = "Month:"
    ws["D3"] = "June"
    start = 5
    for i, row in enumerate(KPI_ROWS):
        for j, val in enumerate(row):
            ws.cell(row=start + i, column=1 + j, value=val)
    return wb


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--module", choices=["kitchen", "farm", "facilities", "all"], default="all")
    args = p.parse_args()
    OUT.mkdir(parents=True, exist_ok=True)
    modules = (
        ["kitchen", "farm", "facilities"] if args.module == "all" else [args.module]
    )
    titles = {
        "kitchen": "Kitchen Operations Dashboard",
        "farm": "Farm Operations Dashboard",
        "facilities": "Facilities Operations Dashboard",
    }
    for m in modules:
        path = OUT / f"{m}_master_template.xlsx"
        build_workbook(titles[m]).save(path)
        print(f"Created {path}")
    print("Upload via Admin → Load data → select module.")


if __name__ == "__main__":
    main()
