#!/usr/bin/env python3
"""
Extract every tab from a Google Spreadsheet to CSV and JSON.

Auth (pick one):
  1. Service account JSON — set GOOGLE_APPLICATION_CREDENTIALS or pass --credentials
     Share the sheet with the service account email (Viewer is enough).
  2. Local .xlsx — File → Download → Microsoft Excel in Google Sheets, then:
     python scripts/extract_google_sheet.py --xlsx path/to/workbook.xlsx

Public sheet (no auth): add --public if the sheet is "Anyone with the link can view".
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any


def json_default(obj: Any) -> Any:
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, time):
        return obj.isoformat()
    raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable")

SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
DEFAULT_SPREADSHEET_ID = "1BDkvHWhJnJXS9vx1c2reyXkab494ZF7bji8z76Ck-mw"


def parse_spreadsheet_id(value: str) -> str:
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", value)
    return m.group(1) if m else value.strip()


def slugify(name: str) -> str:
    s = re.sub(r"[^\w\s-]", "", name, flags=re.UNICODE)
    s = re.sub(r"[-\s]+", "_", s.strip())
    return s[:80] or "sheet"


def write_csv(path: Path, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(["" if c is None else c for c in row])


def write_json(path: Path, title: str, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = [str(h) if h is not None else "" for h in (rows[0] if rows else [])]
    data_rows = rows[1:] if len(rows) > 1 else []
    records = []
    for row in data_rows:
        rec = {}
        for i, h in enumerate(headers):
            key = h or f"col_{i}"
            rec[key] = row[i] if i < len(row) else None
        records.append(rec)
    payload = {"sheet": title, "headers": headers, "rows": records, "raw": rows}
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=json_default),
        encoding="utf-8",
    )


def extract_xlsx(xlsx_path: Path, out_dir: Path) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    written: list[str] = []
    for ws in wb.worksheets:
        rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        slug = slugify(ws.title)
        write_csv(out_dir / f"{slug}.csv", rows)
        write_json(out_dir / f"{slug}.json", ws.title, rows)
        written.append(ws.title)
    wb.close()
    return written


def _build_service(credentials_path: str | None):
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    creds_file = credentials_path or __import__("os").environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if not creds_file or not Path(creds_file).is_file():
        raise SystemExit(
            "No credentials found. Set GOOGLE_APPLICATION_CREDENTIALS or pass --credentials PATH "
            "to a service account JSON file, and share the sheet with that account email."
        )
    creds = service_account.Credentials.from_service_account_file(creds_file, scopes=SCOPES)
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def extract_api(
    spreadsheet_id: str,
    out_dir: Path,
    credentials_path: str | None,
) -> list[str]:
    service = _build_service(credentials_path)
    meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheets = meta.get("sheets", [])
    written: list[str] = []

    for sheet in sheets:
        props = sheet["properties"]
        title = props["title"]
        slug = slugify(title)
        result = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=f"'{title}'")
            .execute()
        )
        rows = result.get("values", [])
        write_csv(out_dir / f"{slug}.csv", rows)
        write_json(out_dir / f"{slug}.json", title, rows)
        written.append(title)

    return written


def extract_public(spreadsheet_id: str, out_dir: Path) -> list[str]:
    import urllib.request

    # Published HTML embed lists sheet gids; fallback: export whole workbook as xlsx if allowed
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    dest = out_dir / "_workbook.xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            dest.write_bytes(resp.read())
    except Exception as e:
        raise SystemExit(
            f"Public export failed ({e}). Share the sheet as 'Anyone with the link' or use API credentials."
        ) from e
    titles = extract_xlsx(dest, out_dir)
    dest.unlink(missing_ok=True)
    return titles


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract all tabs from a Google Sheet")
    parser.add_argument(
        "--spreadsheet-id",
        default=DEFAULT_SPREADSHEET_ID,
        help="Spreadsheet ID or full Google Sheets URL",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "data" / "sheets",
        help="Output directory (default: data/sheets)",
    )
    parser.add_argument("--credentials", help="Path to Google service account JSON")
    parser.add_argument("--xlsx", type=Path, help="Local Excel export (skips API)")
    parser.add_argument(
        "--public",
        action="store_true",
        help="Download via public export URL (sheet must be link-viewable)",
    )
    args = parser.parse_args()

    out_dir = args.output.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.xlsx:
        if not args.xlsx.is_file():
            sys.exit(f"File not found: {args.xlsx}")
        titles = extract_xlsx(args.xlsx.resolve(), out_dir)
        mode = "xlsx"
    elif args.public:
        sid = parse_spreadsheet_id(args.spreadsheet_id)
        titles = extract_public(sid, out_dir)
        mode = "public"
    else:
        sid = parse_spreadsheet_id(args.spreadsheet_id)
        titles = extract_api(sid, out_dir, args.credentials)
        mode = "api"

    manifest = {
        "spreadsheet_id": parse_spreadsheet_id(args.spreadsheet_id),
        "mode": mode,
        "output_dir": str(out_dir),
        "tabs": titles,
        "count": len(titles),
    }
    (out_dir / "manifest.json").write_text(
        json.dumps(manifest, indent=2), encoding="utf-8"
    )
    print(f"Extracted {len(titles)} tab(s) to {out_dir}")
    for t in titles:
        print(f"  - {t}")


if __name__ == "__main__":
    main()
